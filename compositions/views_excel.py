from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from io import BytesIO
import math


@api_view(['POST'])
@permission_classes([AllowAny])
def export_excel(request):
    print("Request Data: ", request.data)  # Debug line
    items = request.data.get('items', [])
    BDI = request.data.get('BDI', 0.1)  # Default value can be set here if needed
    name = request.data.get('name', '')  # Default value can be set here if needed
    desonerado = request.data.get('desonerado', 'nao_desonerado')  # Default value can be set here if needed
    print("BDI Value: ", BDI)  # Debug line
    

    wb = Workbook()
    ws = wb.active

    arial_font = Font(name='Arial', size=10)
    number_style = NamedStyle(name='number_style', number_format="#,##0.00")

    gray_border = Border(left=Side(style='thin', color='808080'),
                     right=Side(style='thin', color='808080'),
                     top=Side(style='thin', color='808080'),
                     bottom=Side(style='thin', color='808080'))

   # These are the sub-headers
    sub_headers = ['Item', 'Código', 'Descrição', 'Unid', 'Qtd', 'Custo Unit', 'MO', 'Material', 'Total', 'MO', 'Material', 'Total']


    for col_num, header in enumerate(sub_headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}4'.format(col_letter)] = header
        ws['{}4'.format(col_letter)].alignment = Alignment(horizontal='center')

    # Merge and set overarching header for "Preço Unitário"
    ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=9)
    ws['G4'].value = 'Preço Unitário'
    ws['G4'].alignment = Alignment(horizontal='center')
    ws['G4'].font = Font(bold=True)

    # Merge and set overarching header for "Total"
    ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=12)
    ws['J4'].value = 'Total'
    ws['J4'].alignment = Alignment(horizontal='center')
    ws['J4'].font = Font(bold=True)


    # Inserting the subheaders for columns 1 to 6 in row 4
    for col_num in [1, 2, 3, 4, 5, 6]:
        col_letter = get_column_letter(col_num)
        ws.merge_cells(start_row=4, start_column=col_num, end_row=5, end_column=col_num)
        ws['{}4'.format(col_letter)].value = sub_headers[col_num - 1]
        ws['{}4'.format(col_letter)].alignment = Alignment(horizontal='center')
        ws['{}4'.format(col_letter)].font = Font(bold=True)

    # Inserting the subheaders for columns 7 to 12 in row 5
    for col_num in range(7, 13):
        col_letter = get_column_letter(col_num)
        ws['{}5'.format(col_letter)] = sub_headers[col_num - 1]
        ws['{}5'.format(col_letter)].alignment = Alignment(horizontal='center')
        ws['{}5'.format(col_letter)].font = Font(bold=True)


    # Styling
    for row_num in range(1, 4):
        ws.row_dimensions[row_num].height = 25

    ws.column_dimensions['C'].width = 64
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['L'].width = 10
    
    row_num = 6  # Start adding budget data from row 5

    bold_font = Font(bold=True, name='Arial')
    center_alignment = Alignment(horizontal='center', vertical='center')


    # Additional headers and meta-info
    ws['C1'] = 'Obra'
    ws['C1'].font = Font(name='Arial', size=10, bold=True)
    ws['C2'] = name

    ws['E1'] = 'Bancos'
    ws['E1'].font = Font(name='Arial', size=10, bold=True)
    ws['E2'] = 'SINAPI - 08/2023'

    ws['H1'] = 'B.D.I.'
    ws['H1'].font = Font(name='Arial', size=10, bold=True)
    BDI_percentage = BDI * 100
    ws['H2'] = f"{BDI_percentage}%"

    ws['K1'] = 'Encargos Sociais'
    ws['K2'].font = Font(name='Arial', size=10, bold=True)
    ws['K2'] = 'Não Desonerado'

    # Merge and Center Cell for Row 3
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(sub_headers))
    centered = Alignment(horizontal='center', vertical='center')
    ws['A3'].alignment = centered
    ws['A3'].value = "Planilha Orçamentária"
    ws['A3'].font = Font(name='Arial', size=10, bold=True)

    # Set height for row 2
    ws.row_dimensions[2].height = 60


    for item in items:

        if item.get('type') == 'stage':
            # Only include 'refId' and 'name' for stage rows
            refId = item.get('refId', '').upper()
            name = item.get('name', '').upper()
            stagetotal = item.get('totalCost', 0)
            row = [refId, '', name, '', '', '', '', '', '', '', '', stagetotal]
            ws.row_dimensions[row_num].height = 30
            for i in range(len(row), len(sub_headers)):
                row.append('')  # Fill the remaining columns with empty strings

        else:
            refId = item.get('refId', '')
            codigo = item.get('codigo', '')
            name = item.get('name', '')
            unit = item.get('unit', '')
            quantity = item.get('quantity', 0)
            total_cost = item.get('total_cost', 0)
            mo_cost = item.get('mo_cost', 0)
            material_cost = item.get('material_cost', 0)
            costWithBDI = item.get('costWithBDI', 0)

            preco_unit_mo = math.floor(mo_cost * (1 + BDI) * 100) / 100.0
            preco_unit_material = costWithBDI - preco_unit_mo
            preco_unit_total = costWithBDI
            total_mo = preco_unit_mo * quantity
            total_material = preco_unit_material * quantity
            total = preco_unit_total * quantity

            row = [refId, codigo, name, unit, quantity, total_cost, preco_unit_mo, preco_unit_material, preco_unit_total, total_mo, total_material, total]
            ws.row_dimensions[row_num].height = 45
            
        ws.append(row)
        print("Appending row:", row)

        if item.get('type') == 'stage':
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=11)
            ws.cell(row=row_num, column=3).value = name  # Set the value for the merged cell


        for col_num in range(1, len(sub_headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = arial_font
            cell.alignment = Alignment(wrap_text=True)
            cell.border = gray_border
            
            # Center-align all columns except 'Name/Description'
            if col_num != 3:
                cell.alignment = center_alignment

            if col_num in [6, 7, 8, 9, 10, 11, 12]:
                 cell.number_format = "#,##0.00"


     

         # Wrap text for the 3rd column (Name/Description)
        cell_to_wrap = ws.cell(row=row_num, column=3)
        cell_to_wrap.alignment = Alignment(wrap_text=True)

        if item.get('type') == 'stage':
            fill = PatternFill(start_color='DDEBF7',
                               end_color='DDEBF7', fill_type='solid')
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = fill
                cell.font = bold_font  # Make the text bold for stage rows
        
        row_num += 1

    # Save it to a BytesIO object
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HttpResponse
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Budget.xlsx'
    return response